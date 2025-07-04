# Importing load_workbook to work with Excel files using openpyxl
from openpyxl import load_workbook 

# Function to read test data (username, password, expected result) from a given Excel file and sheet
def get_test_data(path, sheet_name):
    # Load the Excel workbook from the given file path
    workbook = load_workbook(path)
    # Select the specific sheet by name
    sheet = workbook[sheet_name]
    # Initialize an empty list to hold the data rows
    data = []

    # Iterate over each row in the sheet starting from the second row (excluding headers), up to column 8
    for row in sheet.iter_rows(min_row=2, max_col=8, values_only=True):
        # Extract username from column F (index 5)
        username = row[5]
        # Extract password from column G (index 6)
        password = row[6]
        # Extract expected result (valid/invalid) from column H (index 7)
        expected_result = row[7]
        # Append the data as a tuple to the list
        data.append((username, password, expected_result))
    
    # Return the complete list of data rows
    return data

# Function to find the row number in Excel that matches a given username
def find_row_number(username, path="data/test_data.xlsx", sheet_name="Sheet1"):
    # Load the Excel workbook from the given file path
    workbook = load_workbook(path)
    # Select the specific sheet by name
    sheet = workbook[sheet_name]

    # Enumerate through each row starting from row 2, collecting index (starting from 2)
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # If the username in column F (index 5) matches the input username, return its row number
        if row[5] == username:
            return idx
    # Return -1 if no matching username is found
    return -1  # Not found
