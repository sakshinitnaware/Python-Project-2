# Import the load_workbook function from the openpyxl library to read/write Excel files
from openpyxl import load_workbook 


# Define a function to write the test result (PASS/FAIL) into a specific cell in an Excel sheet
def write_test_result(path, sheet_name, row_num, result):
    # Load the workbook (Excel file) from the given path
    workbook = load_workbook(path)

    # Select the specific sheet by name
    sheet = workbook[sheet_name]

    # Write the result into column 8 (H column) of the specified row
    sheet.cell(row=row_num, column=8).value = result  # Column H = 8

    # Save the updated workbook back to the same file
    workbook.save(path)
